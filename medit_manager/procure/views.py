from datetime import date, datetime
from io import BytesIO

import openpyxl
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView, ListView, DetailView, UpdateView, DeleteView, TemplateView
)

from .forms import ProcurementRecordForm, PurchasedItemFormSet
from .models import ProcurementRecord, ProcurementAttachment, PurchasedItem


class ProcurementRecordListView(ListView):
    model = ProcurementRecord
    template_name = "procure/record_list.html"
    context_object_name = "records"
    paginate_by = 10

    def get_queryset(self):
        qs = ProcurementRecord.objects.all()

        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(title__icontains=q))

        month = self.request.GET.get("month")
        year = self.request.GET.get("year")
        # Lọc theo date_recorded (ngày do người dùng nhập)
        if month and year:
            qs = qs.filter(date_recorded__month=month, date_recorded__year=year)

        is_paid = self.request.GET.get("is_paid")
        if is_paid in ["0", "1"]:
            qs = qs.filter(is_paid=bool(int(is_paid)))

        return qs.order_by("-date_recorded")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        records = context.get("records", [])

        # Gắn tổng tiền tính động (nếu muốn hiển thị nhanh)
        for record in records:
            # đảm bảo total_cost được cập nhật nếu bạn dùng calc_total_cost để lưu
            try:
                record.calculated_total_cost = record.calc_total_cost()
            except Exception:
                record.calculated_total_cost = getattr(record, "total_cost", 0)

        context["q"] = self.request.GET.get("q", "")
        context["month"] = self.request.GET.get("month", "")
        now = datetime.now()
        context["years"] = list(range(2020, now.year + 1))
        context["current_year"] = now.year
        context["months"] = [f"{i:02d}" for i in range(1, 13)]
        return context


class ProcurementRecordDetailView(DetailView):
    model = ProcurementRecord
    template_name = "procure/record_detail.html"
    context_object_name = "record"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Tính tổng tiền động cho hiển thị (không bắt buộc là giá trong DB)
        context["total_cost"] = self.object.calc_total_cost()
        return context


class ProcurementRecordCreateView(CreateView):
    model = ProcurementRecord
    form_class = ProcurementRecordForm
    template_name = "procure/record_form.html"
    success_url = reverse_lazy('procure:record_list')

    def get_initial(self):
        initial = super().get_initial()
        initial['date_recorded'] = date.today()
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.method == "POST":
            context['formset'] = PurchasedItemFormSet(self.request.POST, self.request.FILES)
        else:
            context['formset'] = PurchasedItemFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context.get('formset')

        if formset is None:
            return self.form_invalid(form)

        if not formset.is_valid():
            return self.form_invalid(form)

        # 1) Lưu hồ sơ chính trước (bao gồm date_recorded, is_paid, ...)
        self.object = form.save()

        # 2) Lưu attachments (nếu có)
        attachments = form.cleaned_data.get('attachments')
        if attachments:
            # attachments có thể là list do custom MultipleFileField
            if isinstance(attachments, (list, tuple)):
                for f in attachments:
                    ProcurementAttachment.objects.create(record=self.object, file=f)
            else:
                ProcurementAttachment.objects.create(record=self.object, file=attachments)

        # 3) Lưu items từ formset (commit=False để gán record trước khi save)
        items = formset.save(commit=False)
        for item in items:
            item.record = self.object
            item.save()
        # Xóa những item bị đánh dấu DELETE
        for obj in formset.deleted_objects:
            obj.delete()
        # gọi formset.save() để sync M2M/internal state nếu cần
        formset.save()

        # 4) Nếu có file Excel import, đọc và tạo thêm item
        excel_file = self.request.FILES.get('import_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()))
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # cấu trúc file: STT | Tên | Đơn vị | Số lượng | Đơn giá | (opt) Thanh toán
                    # Bỏ cột STT nếu có
                    if not row:
                        continue
                    # cố gắng lấy các cột đúng thứ tự
                    name = row[1] if len(row) > 1 else None
                    unit = row[2] if len(row) > 2 else ''
                    quantity = row[3] if len(row) > 3 else 0
                    unit_price = row[4] if len(row) > 4 else 0
                    if name:
                        PurchasedItem.objects.create(
                            record=self.object,
                            name=str(name).strip(),
                            unit=unit or '',
                            quantity=int(quantity or 0),
                            unit_price=unit_price or 0
                        )
            except Exception as e:
                messages.warning(self.request, f"Lỗi khi đọc Excel: {e}")

        # 5) Cập nhật tổng tiền chính xác và lưu vào DB
        try:
            self.object.calc_total_cost()  # đảm bảo calc_total_cost() lưu total_cost nếu bạn đã triển khai lưu trong model
        except Exception:
            # fallback: tính tạm thời rồi lưu
            total = sum(item.total_price() for item in self.object.items.all())
            self.object.total_cost = total
            self.object.save(update_fields=['total_cost'])

        return super().form_valid(form)


class ProcurementRecordUpdateView(UpdateView):
    model = ProcurementRecord
    form_class = ProcurementRecordForm
    template_name = "procure/edit_record.html"
    success_url = reverse_lazy('procure:record_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.method == "POST":
            context['formset'] = PurchasedItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = PurchasedItemFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context.get('formset')

        if formset is None:
            return self.form_invalid(form)

        if not formset.is_valid():
            return self.form_invalid(form)

        # Lưu record chính
        self.object = form.save()

        # Lưu attachments mới (nếu có)
        attachments = form.cleaned_data.get('attachments')
        if attachments:
            if isinstance(attachments, (list, tuple)):
                for f in attachments:
                    ProcurementAttachment.objects.create(record=self.object, file=f)
            else:
                ProcurementAttachment.objects.create(record=self.object, file=attachments)

        # Lưu items formset
        items = formset.save(commit=False)
        for item in items:
            item.record = self.object
            item.save()
        # Xóa item đã đánh dấu
        for obj in formset.deleted_objects:
            obj.delete()
        formset.save()

        # Cập nhật tổng tiền
        try:
            self.object.calc_total_cost()
        except Exception:
            total = sum(item.total_price() for item in self.object.items.all())
            self.object.total_cost = total
            self.object.save(update_fields=['total_cost'])

        return super().form_valid(form)


class ProcurementRecordDeleteView(DeleteView):
    model = ProcurementRecord
    template_name = "procure/delete_record.html"
    success_url = reverse_lazy('procure:record_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['record'] = self.object
        return context


class DashboardView(TemplateView):
    template_name = "procure/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["total_records"] = ProcurementRecord.objects.count()
        context["total_paid"] = ProcurementRecord.objects.filter(is_paid=True).aggregate(Sum("total_cost"))["total_cost__sum"] or 0
        context["total_unpaid"] = ProcurementRecord.objects.filter(is_paid=False).aggregate(Sum("total_cost"))["total_cost__sum"] or 0

        # Thống kê theo tháng dựa trên date_recorded
        current_year = datetime.now().year
        monthly_data = (
            ProcurementRecord.objects
            .filter(date_recorded__year=current_year)
            .values("date_recorded__month")
            .annotate(total=Sum("total_cost"))
            .order_by("date_recorded__month")
        )

        monthly_totals = {m: 0 for m in range(1, 13)}
        for entry in monthly_data:
            month_key = entry.get("date_recorded__month")
            total = entry.get("total") or 0
            monthly_totals[month_key] = total

        context["monthly_totals"] = monthly_totals
        context["current_year"] = current_year

        # Debug print (bạn có thể xóa sau khi kiểm tra)
        print("Monthly Totals:", monthly_totals)

        return context


def export_excel(request):
    qs = ProcurementRecord.objects.all()
    q = request.GET.get("q")
    month = request.GET.get("month")
    year = request.GET.get("year")

    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(title__icontains=q))
    if month and year:
        qs = qs.filter(date_recorded__month=month, date_recorded__year=year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách hồ sơ"

    ws.append(["Mã hồ sơ", "Tiêu đề", "Tổng trị giá", "Ngày lập hồ sơ"])
    for r in qs:
        # dùng total_cost lưu trong DB (đã được cập nhật khi lưu hồ sơ)
        ws.append([r.code, r.title, float(r.total_cost or 0), str(r.date_recorded)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="hoso_muasam.xlsx"'
    wb.save(response)
    return response
